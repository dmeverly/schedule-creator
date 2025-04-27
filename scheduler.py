import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
import calendar
import math

INPUT = "Template.xlsx"
OUTPUT = "Schedule.xlsx"
DOW = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
numMonths = 12

def readXlsx():
    workbook = openpyxl.load_workbook(INPUT)
    workspace = workbook.active
    dataFrame = list(workspace.iter_rows(values_only=True))
    df = pd.DataFrame(data=dataFrame, dtype=str)
    return df

def createSheet(d1, d2, n, weekNumber, month, year, Styles, wb):
    title = f"{calendar.month_name[month]} {year}"
    ws = wb.create_sheet(title=title)

    ws["A1"] = title
    ws["A1"].font = Styles.get('month')
    ws["A1"].alignment = Styles.get('center')
    ws["A1"].border = Styles.get('thinb')
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=14)

    for i, day in enumerate(DOW):
        col = i * 2 + 1
        ws.cell(row=3, column=col).value = day
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws.cell(row=3, column=col).font = Styles.get('day')
        ws.cell(row=3, column=col).alignment = Styles.get('center')

    #day indexed 0-6 Monday-Sunday
    monthStartDay, monthLength = calendar.monthrange(year, month)
    #add one day so index 0-6 Sunday-Saturday
    monthStartDay += 1
    if monthStartDay > 6:
        monthStartDay = 0
    weeksInMonth = math.ceil((monthStartDay + monthLength) / 7)
    startColumn = (monthStartDay * 2) + 1
    date = 1
    templateDay = monthStartDay
    alternate = 0

    for week in range(weeksInMonth):
        for day in range(7):
            col = day * 2 + 1
            if week == 0 and col < startColumn:
                continue
            if date > monthLength:
                break

            d1_emp = d1[templateDay][weekNumber]
            d2_emp = d2[templateDay][weekNumber]
            n_emp = n[templateDay][weekNumber]
 
            shift_text = f"{date}\n{d1_emp} - Day\n{d2_emp} - Day\n{n_emp} - Night"
            if day == 6:
                shift_text += f"\nTemplate Week {weekNumber}"
                shift_text += " - Dr. Amin" if alternate == 0 else ""
                alternate = 1 - alternate
            
            cell = ws.cell(row=4 + week, column=col, value=shift_text)
            cell.alignment = Styles.get('tla')
            cell.font = Styles.get('cell')
            ws.merge_cells(start_row=4 + week, start_column=col, end_row=4 + week, end_column=col + 1)
            
            for merge_col in range(col, col + 2):
                ws.cell(row=4 + week, column=merge_col).border = Styles.get('thinb')

            date += 1
            templateDay += 1
            if templateDay > 6:
                templateDay = 0
                weekNumber += 1
                if weekNumber > 12:
                    weekNumber = 1

    for row in range(3, 4 + weeksInMonth):
        for col in range(1, 15):
            ws.cell(row=row, column=col).border = Styles.get('thinb')

    for col in range(1, 15):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    for row in range(4, 4 + weeksInMonth):
        max_lines = max(
            ws.cell(row=row, column=col).value.count("\n") + 1 
            if ws.cell(row=row, column=col).value else 1
            for col in range(1, 14, 2)
        )
        ws.row_dimensions[row].height = max(25, max_lines * 13)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1

    return weekNumber

def preProcess(df):
    df = df.drop([0, 1, 14, 15])
    df = df.drop([0, 8, 16], axis=1)
    df.columns = list(range(21))
    df.index = list(range(1, 13))

    #days index 0-6 sunday-saturday, weeks index 1-12
    d1 = df.iloc[:, :7]
    d1.columns = list(range(7))
    d2 = df.iloc[:, 7:14]
    d2.columns = list(range(7))
    n = df.iloc[:, 14:]
    n.columns = list(range(7))
    
    return d1, d2, n

def queryInput(string, type):
    keyboard = input(string)
    valid = False
    match type:
        case 'week':
            while(not valid):
                try:
                    value = int(keyboard)
                    if value > 0 and value < 13:
                        valid = True
                except:
                    print("Invalid Entry")
        case 'month':
            while(not valid):
                try:
                    value = int(keyboard)
                    if value > 0 and value < 13:
                        valid = True
                except:
                    print("Invalid Entry")
        case 'year':
            while(not valid):
                try:
                    value = int(keyboard)
                    if value > 2024 and value < 2030:
                        valid = True
                except:
                    print("Invalid Entry")
    return value

if __name__ == "__main__":
    d1, d2, n = preProcess(readXlsx())

    weekNumber = queryInput("Template Week Number: ", "week")
    monthStart = queryInput("Starting Month Number: ", "month")
    year = queryInput("Starting Year: ", "year")

    Styles = {
        'month': Font(name="Arial", size=14, bold=True),
        'day': Font(name="Arial", size=11, bold=True),
        'cell': Font(name="Arial", size=10),
        'tla': Alignment(horizontal="left", vertical="top", wrap_text=True),
        'center': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'thinb': Border(bottom=Side(style="thin"), top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    }

    wb = Workbook()
    month = monthStart
    for i in range(numMonths):
        if month == 13:
            month = 1
            year += 1
        weekNumber = createSheet(d1, d2, n, weekNumber, month, year, Styles, wb)
        month += 1

    wb.remove(wb['Sheet'])
    wb.save(OUTPUT)
    print(f"Schedule saved as {OUTPUT}")
